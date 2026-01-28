import pandas as pd
from openpyxl import load_workbook


def load_file(path):
    if path.endswith(".csv"):
        return pd.read_csv(path, header=None)

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    df = pd.DataFrame(data)

    for merged in ws.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged.bounds
        value = ws.cell(min_row, min_col).value
        for r in range(min_row - 1, max_row):
            for c in range(min_col - 1, max_col):
                df.iat[r, c] = value

    return df
